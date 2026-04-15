import json
import time
import random
import os
import re
import requests
import pandas as pd
from playwright.sync_api import sync_playwright
from datetime import datetime
import easyocr
import warnings
import sys
import argparse
import glob
import html
import openpyxl
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
import torch

# THE SILENCER: Mengurangi noise terminal
warnings.filterwarnings("ignore", category=UserWarning, module="torch.utils.data.dataloader")

class BPSMultimodalScraper:
    def __init__(self, config_path="config/targets.json", auth_path="config/auth_state.json", reset_state=False):
        self._initialize_config(config_path)
        self.auth_path = auth_path
        self.base_data_path = "data/audit_results/"
        self.raw_output_path = "data/raw/"
        self.progress_file = "config/progress.json"
        
        self.cleanup_old_media(days_old=30)

        self.intercepted_dates = {}
        self.intercepted_captions = {} 
        
        if reset_state:
            self._reset_progress()
            
        self.completed_targets = self._load_progress()
        self.current_session_results = []
        os.makedirs(self.base_data_path, exist_ok=True)
        os.makedirs(self.raw_output_path, exist_ok=True)
        
        print("[SYSTEM] Mengaktifkan Neural Engine (EasyOCR)...")
        has_gpu = torch.cuda.is_available()
        if has_gpu:
            print("[SYSTEM] 🟢 GPU Terdeteksi (CUDA)! Kecepatan ekstraksi maksimal aktif.")
            self.reader = easyocr.Reader(['id', 'en'], gpu=True)
        else:
            print("[SYSTEM] 🟡 GPU Tidak Ditemukan. Jatuh kembali ke mode CPU.")
            print("         *Reminder: Jika ini laptop pribadi, proses per-gambar akan sedikit lebih lambat.")
            self.reader = easyocr.Reader(['id', 'en'], gpu=False)
                    
        print("[SYSTEM] Neural Engine Siap.")

        # =====================================================================
        # EXPANDED TAXONOMY: Pemetaan Sosiolinguistik Politik & Filantropi
        # =====================================================================
        self.keywords = {
            "realisasi_pengeluaran_sosial": [
                r"\btersalurkan\b", r"\brealisasi\b", r"\bdistribusi\b", r"\bpenyaluran\b", 
                r"\bpenyerahan sumbangan\b", r"\bmenyerahkan bantuan\b", r"\bpaket makanan\b",
                r"\bpangan yatim\b", r"\bbantuan operasional\b", r"\bsumbangan dana\b"
            ],
            "kampanye_fundraising": [
                r"\byuk donasi\b", r"\bscan qris\b", r"\bsalurkan donasi\b", r"\btransfer ke\b",
                r"\bopen donasi\b", r"\brekening donasi\b", r"\btunaikan zakat\b", r"\bbayar zakat\b",
                r"\bsedekah\b", r"\binfaq\b", r"\bgalang dana\b", r"\bklik link\b"
            ],
            "advokasi_dan_pekerja": [
                r"\bpetisi\b", r"\bsikap\b", r"\bregulasi\b", r"\btolak\b", r"\bserikat pekerja\b",
                r"\brealita miris\b", r"\beksploitasi\b", r"\bdimanipulasi\b", r"\bkesejahteraan\b",
                r"\bnasib para\b", r"\bkebutuhan hidup layak\b", r"\baudiensi\b", r"\btuntutan\b"
            ],
            "konsolidasi_dan_politik": [
                r"\bverifikasi faktual\b", r"\bkpu\b", r"\bbawaslu\b", r"\bsilaturahmi\b",
                r"\bpenyerahan sk\b", r"\bpelantikan pengurus\b", r"\bpimda\b", r"\bdpc\b",
                r"\brakercab\b", r"\bkonsolidasi\b", r"\bsambangi\b", r"\bwali kota\b"
            ],
            "edukasi_religi_dan_hari_besar": [
                r"\bhukum zakat\b", r"\bkeutamaan\b", r"\bpahala\b", r"\bdalil\b",
                r"\bmarhaban ya ramadan\b", r"\bibadah puasa\b", r"\bidul fitri\b",
                r"\bhewan qurban\b", r"\bpenyembelihan\b", r"\bhikmah\b", r"\bberkah\b", r"\bsyariat\b"
            ],
            "simpati_dan_duka_cita": [
                r"\bturut berduka cita\b", r"\bberpulangnya ke rahmatullah\b", r"\bwafatnya\b",
                r"\bamal ibadahnya diterima\b", r"\bkeluarga yang ditinggalkan\b", r"\bketabahan\b",
                r"\bhusnul khotimah\b"
            ]
        }
        # =====================================================================

    def _initialize_config(self, path):
        if not os.path.exists(path):
            print(f"[!] Critical: {path} tidak ditemukan."); sys.exit(1)
        with open(path, 'r') as f: self.config = json.load(f)

    def cleanup_old_media(self, days_old=30):
        """Menghapus gambar bukti yang usianya melebihi batas hari (Garbage Collection)."""
        print(f"\n[SYSTEM] Memulai Pembersihan Media (Retention: {days_old} hari)...")
        now = time.time()
        deleted_count = 0
        
        for root, dirs, files in os.walk(self.base_data_path):
            if "media" in root:
                for file in files:
                    if file.endswith(('.jpg', '.png', '.jpeg')):
                        filepath = os.path.join(root, file)
                        # Hitung umur file dalam hitungan hari
                        file_age_days = (now - os.path.getmtime(filepath)) / (24 * 3600)
                        
                        if file_age_days > days_old:
                            try:
                                os.remove(filepath)
                                deleted_count += 1
                            except Exception as e:
                                pass
                                
        if deleted_count > 0:
            print(f"[SYSTEM] Berhasil membersihkan {deleted_count} file gambar usang.")
        else:
            print("[SYSTEM] Ruang penyimpanan media masih bersih.")
    
    def _reset_progress(self):
        print("\n[SYSTEM] Protokol Reset Diaktifkan: Menghapus rekam jejak audit sebelumnya...")
        if os.path.exists(self.progress_file): 
            os.remove(self.progress_file)
        self.completed_targets = []
        old_files = glob.glob(os.path.join(self.base_data_path, "*/*.xlsx"))
        for f in old_files:
            try: os.remove(f)
            except: pass
        print("[SYSTEM] Data Excel lama berhasil dibersihkan.")

    def _load_progress(self):
        if os.path.exists(self.progress_file):
            with open(self.progress_file, 'r') as f: return json.load(f).get("completed", [])
        return []

    def _save_progress(self, username):
        if username not in self.completed_targets:
            self.completed_targets.append(username)
            with open(self.progress_file, 'w') as f: json.dump({"completed": self.completed_targets}, f)

    def clean_text(self, text):
        text = re.sub(r'[\\/*_\[\]{}|~^`]', '', text)
        return " ".join(text.replace('Bartai', 'Partai').split())

    def auto_flagger(self, text):
        text = str(text).lower()
        flags = [cat for cat, pats in self.keywords.items() if any(re.search(p, text) for p in pats)]
        return ", ".join(flags) if flags else "konten_umum"

    def extract_highest_nominal(self, text):
        """
        Semantic & Resilient Financial Parser v2.0
        Tahan terhadap typo OCR, kebal terhadap nomor HP/Rekening, dan presisi tinggi.
        """
        if not text or text == "OCR_FAILED": 
            return 0
        
        # 1. PRE-PROCESSING
        text_clean = str(text).lower()
        text_clean = re.sub(r'(?<=\d)[sS](?=[.,\d]|$)|(?<=^[.,\d])[sS](?=\d)', '5', text_clean)
        text_clean = re.sub(r'(?<=\d)[bB](?=[.,\d]|$)|(?<=^[.,\d])[bB](?=\d)', '8', text_clean)
        text_clean = re.sub(r'(?<=\d)[oO]+|[oO]+(?=\d)', '0', text_clean)
        # Bersihkan noise pemisah (misal: ,= atau -,)
        text_clean = text_clean.replace(',=', '').replace('-,', '.')
        
        nominals = []

        # 2. EKSTRAKSI FORMAT RUPIAH (Prefix Diperketat, Angka Diperlonggar)
        # Menghapus huruf tunggal 'r' dan 'p'. 
        # [.\-\s]* memungkinkan penangkapan typo seperti "Pp-25000" atau "Rp. 25000"
        pola_rp = r'\b(?:rp|pp|bp|idr)[.\-\s]*([\d]+(?:[.,]\d+)*)'
        for match in re.findall(pola_rp, text_clean):
            # Hilangkan semua titik dan koma untuk mendapatkan nilai pure integer
            clean_num = re.sub(r'\D', '', match)
            if clean_num.isdigit():
                # Filter 1: Jika diawali '0' dan panjang > 8, itu pasti Nomor HP, bukan uang
                if not (clean_num.startswith('0') and len(clean_num) > 8):
                    nominals.append(int(clean_num))

        # 3. EKSTRAKSI FORMAT NARATIF
        # m dan t sudah dihapus sesuai intuisi analitik Anda yang sangat tepat
        pola_naratif = r'([0-9]+(?:[.,][0-9]+)?)\s?(ribu|rb|juta|jt|miliar|milyar|triliun)\b'
        for angka, skala in re.findall(pola_naratif, text_clean):
            try:
                val = float(angka.replace(',', '.'))
                
                if skala in ['ribu', 'rb']: mult = 1_000
                elif skala in ['juta', 'jt']: mult = 1_000_000
                elif skala in ['miliar', 'milyar']: mult = 1_000_000_000
                elif skala == 'triliun': mult = 1_000_000_000_000
                else: mult = 1
                
                nominals.append(int(val * mult))
            except: 
                continue

        # 4. SANITY CHECK (Batas Akal Sehat)
        # Filter 2: Membuang angka di atas 10 Triliun (untuk mengeleminasi nomor rekening 12-15 digit yang bocor)
        valid_nominals = [n for n in nominals if n < 10000000000000]

        return max(valid_nominals) if valid_nominals else 0

    def _json_crawler(self, data, current_shortcode=None):
        if isinstance(data, dict):
            shortcode = data.get('shortcode') or data.get('code') or current_shortcode
            timestamp = data.get('taken_at_timestamp') or data.get('taken_at')
            if shortcode and isinstance(timestamp, int):
                self.intercepted_dates[shortcode] = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')
            
            caption_text = None
            if 'edge_media_to_caption' in data and isinstance(data['edge_media_to_caption'], dict):
                edges = data['edge_media_to_caption'].get('edges', [])
                if edges and len(edges) > 0: caption_text = edges[0].get('node', {}).get('text')
            elif 'caption' in data and isinstance(data['caption'], dict):
                caption_text = data['caption'].get('text')
                
            if shortcode and caption_text:
                self.intercepted_captions[shortcode] = caption_text

            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    self._json_crawler(value, shortcode)
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, (dict, list)):
                    self._json_crawler(item, current_shortcode)

    def _intercept_network(self, response):
        if response.status_code == 200 and "application/json" in response.headers.get("content-type", ""):
            try:
                data = response.json()
                self._json_crawler(data)
            except Exception: pass

    def extract_post_date_fallback(self, raw_text):
        text = str(raw_text).lower()
        en_months = ["january", "february", "march", "april", "may", "june", "july", "august", "september", "october", "november", "december"]
        for i, m in enumerate(en_months, 1):
            if m in text:
                match = re.search(rf'{m}\s+(\d{{1,2}}),\s+(\d{{4}})', text)
                if match: return f"{match.groups()[1]}-{str(i).zfill(2)}-{match.groups()[0].zfill(2)}"
        id_months = {"januari": "01", "februari": "02", "maret": "03", "april": "04", "mei": "05", "juni": "06", "juli": "07", "agustus": "08", "september": "09", "oktober": "10", "november": "11", "desember": "12"}
        match = re.search(r'(\d{1,2})\s+([a-z]+)\s+(\d{4})', text)
        if match:
            day, month_str, year = match.groups()
            month_num = id_months.get(month_str)
            if not month_num:
                try: month_num = str(en_months.index(month_str) + 1).zfill(2)
                except ValueError: pass
            if month_num: return f"{year}-{month_num}-{day.zfill(2)}"
        return "Tanggal Tidak Diketahui"

    def perform_ocr(self, img_url, post_id, username, entity_path):
        media_folder = os.path.join(entity_path, "media")
        os.makedirs(media_folder, exist_ok=True)
        img_path = os.path.join(media_folder, f"{post_id}.jpg")
        
        try:
            if not os.path.exists(img_path):
                res = requests.get(img_url, timeout=10)
                if res.status_code == 200:
                    with open(img_path, 'wb') as f: f.write(res.content)
            
            result = self.reader.readtext(img_path, detail=0)
            return self.clean_text(" ".join(result)), img_path
        except Exception as e: 
            return f"OCR_FAILED: {str(e)[:20]}", "NO_IMAGE"

    def _harvest_posts(self, page, target_count):
        posts_data = {}
        scrolls = 0
        while len(posts_data) < target_count and scrolls < 15:
            items = page.query_selector_all("a[href*='/p/'], a[href*='/reel/']")
            for item in items:
                link = item.get_attribute('href')
                img = item.query_selector("img")
                if link and img:
                    url = f"https://www.instagram.com{link.split('?')[0]}"
                    if url not in posts_data:
                        raw_alt = img.get_attribute('alt') or ""
                        clean_cap = raw_alt.replace("Photo by", "").replace("on Instagram", "").strip()
                        posts_data[url] = {"src": img.get_attribute('src'), "cap": clean_cap, "raw_alt": raw_alt}
            page.mouse.wheel(0, 1000); time.sleep(2); scrolls += 1
        return posts_data

    def _format_excel_hyperlinks(self, filepath, col_name="Post_URL"):
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value == col_name:
                    col_idx = cell.column
                    break
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"
            wb.save(filepath)
            wb.close()
        except Exception: pass

    def _generate_master_report(self):
        print("\n[SYSTEM] Memulai konsolidasi Master Report...")
        all_files = glob.glob(os.path.join(self.base_data_path, "*/*.xlsx"))
        if not all_files: return None
        
        df_list = []
        for f in all_files:
            try: df_list.append(pd.read_excel(f))
            except Exception as e: print(f"[!] Gagal membaca {f}: {e}")
            
        if df_list:
            master_df = pd.concat(df_list, ignore_index=True)
            if "Image_Evidence" in master_df.columns:
                master_df.drop_duplicates(subset=["Image_Evidence"], keep="last", inplace=True)
            status = "FINAL" if len(self.completed_targets) >= len(self.config['targets']) else "PARTIAL"
            f_name = os.path.join(self.raw_output_path, f"AUDIT_MASTER_{status}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
            
            master_df.to_excel(f_name, index=False, engine='openpyxl')
            self._format_excel_hyperlinks(f_name, "Post_URL")
            return f_name
        return None

    def scrape(self):
        try:
            with sync_playwright() as p:
                headless_cfg = self.config.get('scraper_settings', {}).get('headless_mode', False)
                browser = p.chromium.launch(headless=headless_cfg)
                context = browser.new_context(storage_state=self.auth_path)
                
                for target in self.config['targets']:
                    username = target['username']
                    if username in self.completed_targets:
                        print(f"[SKIP] @{username} sudah tuntas di sesi sebelumnya."); continue
                        
                    print(f"\n[AUDIT] Memproses: @{username}")
                    entity_path = os.path.join(self.base_data_path, username)
                    os.makedirs(entity_path, exist_ok=True)
                    
                    page = None
                    try:
                        page = context.new_page()
                        page.on("response", self._intercept_network)
                        page.goto(f"https://www.instagram.com/{username}/", timeout=60000)
                        try:
                            page.wait_for_selector("main", timeout=15000)
                        except PlaywrightTimeoutError:
                            print(f"\n[!] Gagal memuat profil @{username}.")
                            # Cek apakah dilempar ke halaman login
                            if page.query_selector("input[name='username']"):
                                print("[CRITICAL] Sesi Cookies Habis! Harap perbarui file 'auth_state.json'.")
                                sys.exit(1) # Hentikan skrip karena percuma dilanjut
                            else:
                                print("[WARNING] Elemen 'main' tidak ditemukan. Mungkin terkena Soft-Ban atau koneksi lambat. Melewati akun ini...")
                                continue # Lanjut ke akun berikutnya

                        scripts_text = page.evaluate('''() => {
                            return Array.from(document.querySelectorAll('script[type="application/json"], script[type="application/json+protobuf"]'))
                                        .map(s => s.textContent);
                        }''')
                        for text in scripts_text:
                            try: self._json_crawler(json.loads(text))
                            except: pass
                        
                        max_p = self.config['scraper_settings'].get('max_posts_per_account', 20)
                        raw_data = self._harvest_posts(page, max_p)
                        
                        entity_results = []
                        for idx, (url, data) in enumerate(list(raw_data.items()), 1):
                            post_id = url.rstrip('/').split('/')[-1]
                            print(f"  ├─ Ekstraksi & OCR [{idx}/{len(raw_data)}] | {post_id}")
                            
                            post_date = self.intercepted_dates.get(post_id)
                            if not post_date: post_date = self.extract_post_date_fallback(data['raw_alt'])

                            native_caption = self.intercepted_captions.get(post_id)
                            if not native_caption: native_caption = data['cap']
                            
                            ocr_txt, img_p = self.perform_ocr(data['src'], post_id, username, entity_path)
                            combined = f"CAPTION: {native_caption} | OCR: {ocr_txt}"
                            
                            # LOGIKA BARU: Tidak ada lagi fitur Consistency_Score
                            entity_results.append({
                                "Institution": username,
                                "Detected_Phenomena": self.auto_flagger(combined),
                                "Estimasi_Nilai_Rp": self.extract_highest_nominal(combined),
                                "Tanggal_Postingan": post_date,
                                "Post_URL": url,
                                "Native_Caption": native_caption,
                                "OCR_Fallback": ocr_txt,
                                "Image_Evidence": img_p,
                                "Scraped_Timestamp": datetime.now().isoformat()
                            })
                        
                        excel_name = os.path.join(entity_path, f"{username}_audit.xlsx")
                        pd.DataFrame(entity_results).to_excel(excel_name, index=False, engine='openpyxl')
                        self._format_excel_hyperlinks(excel_name, "Post_URL")
                        
                        self._save_progress(username)
                        print(f"[SUCCESS] Laporan @{username} tuntas.")
                    except Exception as e: 
                        print(f"[ERROR] @{username}: {str(e)[:50]}")
                    finally: 
                        if page: page.close()
                browser.close()
                
        except KeyboardInterrupt: 
            print("\n[INTERRUPT] Sinyal penghentian diterima. Langsung menyusun Master Report...")
        except Exception as e:
            print(f"\n[CRITICAL ERROR] Terjadi kegagalan sistemik: {e}")
        finally: 
            master = self._generate_master_report()
            if master: print(f"[FINAL] Master Audit Konsolidasi berhasil dibuat di: {master}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BPS Multimodal Audit Scraper")
    parser.add_argument('--reset', action='store_true', help='Hapus rekam jejak progress dan mulai audit dari awal')
    args = parser.parse_args()

    scraper = BPSMultimodalScraper(reset_state=args.reset)
    scraper.scrape()